import os
import re
import pandas as pd
import datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def extract_details(file_name, series_folder, series_name=None):
    # Patterns for extracting details from file names
    season_episode_pattern = re.compile(r"S(\d+)\.E(\d+)")
    year_pattern = re.compile(r"\((\d{4})\)")

    # Remove file extension and any leading directory names that might be part of the file name
    file_name_no_ext = os.path.splitext(file_name)[0]

    # If series_folder or series_name is None, replace with an empty string
    series_folder = series_folder or ''
    series_name = series_name or ''

    # Remove the series folder and the series name from the file name
    clean_title = file_name_no_ext.replace(series_folder, '').replace(series_name, '').strip()

    # Extract season and episode numbers
    season_episode_match = season_episode_pattern.search(clean_title)
    if season_episode_match:
        season = int(season_episode_match.group(1))
        episode = int(season_episode_match.group(2))
        clean_title = season_episode_pattern.sub('', clean_title)  # Remove the season/episode from the title
    else:
        season = None
        episode = None

    # Extract year
    year_match = year_pattern.search(clean_title)
    if year_match:
        year = year_match.group(1)
        clean_title = year_pattern.sub('', clean_title)  # Remove the year from the title
    else:
        year = None

    # Clean the title
    clean_title = clean_title.strip(' -.')

    # Return the cleaned title, year, season, and episode
    return clean_title, year, season, episode

def update_excel(media_info, excel_path):
    try:
        existing_df = pd.read_excel(excel_path)
        print(f"Loaded existing data with {len(existing_df)} entries.")
        existing_df['Season'] = pd.to_numeric(existing_df['Season'], errors='coerce').fillna(0).astype(int)
        existing_df['Episode'] = pd.to_numeric(existing_df['Episode'], errors='coerce').fillna(0).astype(int)
    except FileNotFoundError:
        print("Excel file not found, creating a new one.")
        existing_df = pd.DataFrame(columns=["Series", "Title", "Year", "Season", "Episode"])
        
    new_df = pd.DataFrame(media_info)
    new_df.fillna('', inplace=True)  # Replace None with empty string
    new_df['UniqueKey'] = new_df.apply(lambda x: f"{x['Series']} {x['Title']} ({x['Year']})".strip(), axis=1)
    
    combined_df = pd.concat([existing_df, new_df]).drop_duplicates(subset='UniqueKey', keep='first')
    combined_df.sort_values(by=['Series', 'Season', 'Episode'], inplace=True)
    
    combined_df.to_excel(excel_path, index=False)
    print(f"Excel file updated with {len(combined_df)} entries.")

def format_mtime(mtime_value):
    # Converts the mtime value to a human-readable format
    modification_time = datetime.datetime.fromtimestamp(mtime_value)
    return modification_time.strftime('%Y-%m-%d %H:%M:%S')

# Main function to parse media files and update Excel file
def parse_media_files(directory_path, excel_path):
    media_info = []
    existing_files = []

    for subdir, dirs, files in os.walk(directory_path):
        path_parts = subdir.replace(directory_path, '').strip(os.sep).split(os.sep)
        series_name = path_parts[0] if path_parts else None
        season_folder = path_parts[1] if len(path_parts) > 1 else None

        for filename in files:
            if filename.endswith('.mp4'):
                filepath = os.path.join(subdir, filename)
                file_stat = os.stat(filepath)
                clean_title, year, season, episode = extract_details(filename, season_folder, series_name)
                
                if not clean_title:  # If title is empty, use the file's base name minus the extension
                    clean_title = os.path.splitext(filename)[0]
                    
                file_info = {
                    'Series': series_name if series_name else '',
                    'SeasonFolder': season_folder,  # Used for sorting
                    'Title': clean_title,
                    'Year': year if year else '',
                    'Season': season if season is not None else '',
                    'Episode': episode if episode is not None else '',
                    'size': file_stat.st_size,
                    'mtime': format_mtime(file_stat.st_mtime)
                }
                media_info.append(file_info)

    update_excel(media_info, excel_path)

# Path to the media directory and Excel file
media_directory = 'G:/FUTUREMEDIASERVER'
excel_file_path = 'G:/media_library.xlsx'

# Parse the media files and update the Excel file
if __name__ == '__main__':
    parse_media_files(media_directory, excel_file_path)

# Load the workbook and get the active sheet
workbook = load_workbook(excel_file_path)
worksheet = workbook.active

# Define the columns you don't want to auto-fit
exclude_columns = ['path', 'size', 'mtime', 'UniqueKey']

# Get the index of the columns to exclude (if they exist)
exclude_column_indices = []
for col in worksheet[1]:  # Assuming the first row contains headers
    if col.value in exclude_columns:
        exclude_column_indices.append(col.column)

# Iterate through columns and set each one to the width of its longest cell, excluding specified columns
for col_num, column in enumerate(worksheet.columns, start=1):
    if col_num in exclude_column_indices:
        continue  # Skip the columns to exclude
    max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
    adjusted_width = (max_length + 2) * 1.2
    worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

# Save the workbook
workbook.save(excel_file_path)

if 'workbook' in locals():
    workbook.close()